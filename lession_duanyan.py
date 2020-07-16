# -*-coding:utf-8 -*-
#@Time     :2020/7/15
#@Author   :先森
#@Email    :740891494@qq.com
#@File     :lession_duanyan.py
#@Software :PyCharm
#   '5*6'本身是个字符串，可饮用函数eval()把引号去掉，引用引号内的内容


import requests
import openpyxl

#读取测试用例函数
def read_data(filename,sheetname):                      #可变的数据设置为参数，方便往后针对不同的表单读取
    wb = openpyxl.load_workbook(filename)               #加载工作簿——文档名字
    sheet = wb[sheetname]                               #获取表单
    max_row = sheet.max_row                             #获取最大行数
    case_list = []                                      #创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id = sheet.cell(row=i,column=1).value,     #获取case_id
        url = sheet.cell(row=i,column=5).value,         #获取url
        data = sheet.cell(row=i,column=6).value,        #获取data
        expect = sheet.cell(row=i,column=7).value       #获取expect
        )
        case_list.append(dict1)                         #把14条字典存放在列表里面，14条数据.每循环一次就读取到字典数据存放到这个list
        # print(case_list)
    return case_list


#执行接口函数
def api_fun(url,data):

    headers_reg = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}  #请求头——字典
    res = requests.post(url=url,json=data,headers = headers_reg)  #接收post方法的结果
    response = res.json() #相应正文
    return response

#写入结果
def write_result(filename,sheetname,row,column,final_result):    #设置文件名、工作单名、行、列、实际结果形参
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result    #写入结果
    wb.save(filename) #保存，前提是关闭文档




#执行测试用例并回写实际结果
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)         #调用！！#读取文件内的数据，EXCEL表中的register表单

    for case in cases:                                         #一条一条地把数据取出来，放进case
        case_id = case.get('case_id')                          #此时取出来的只是字典，需要提取key对应的值，用到字典取值
        url = case.get('url')
        data = eval(case.get('data'))                          #eval()运行被字符串包裹的表达式——去掉字符串引号
        expect =eval((case.get('expect')) )                         #获取预期结果
        # print(type(data))
        expect_msg = expect.get('msg')                         #获取预期结果中的msg
        real_result = api_fun(url=url,data=data)               #调用！！#调用发送接口请求的函数，返回结果用变量real_result接收
        real_msg = real_result.get('msg')                      #api_fun返回的类型是字典类型，get方法获取实际结果中的msg
        print('预期结果中的msg：{}'.format(expect_msg))
        print('实际结果中的msg：{}'.format(real_msg))
        if real_msg == expect_msg:          #此处可用and来判断多个条件
            print('第{}条测试用例执行通过！'.format(case_id))
            final_re = 'Passed'                                                     #把passed装进final_re
        else:
            print('第{}条测试用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)          #调用！！因为再65行已经有passed装进变量
        print('*'*25)

execute_fun('test_case_api.xlsx','login')